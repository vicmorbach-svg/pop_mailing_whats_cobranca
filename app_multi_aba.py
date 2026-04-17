import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from pathlib import Path

# ── Configuração da página ────────────────────────────────────────────────────
st.set_page_config(
    page_title="Populador de Template Excel por Aba",
    page_icon="📂",
    layout="wide",
)

# ── Caminho do template pré-definido ──────────────────────────────────────────
TEMPLATE_FILE_NAME = "TEMPLATE_WHATS_COBRANCA.xlsx"
TEMPLATE_PATH = Path(__file__).parent / TEMPLATE_FILE_NAME

# ── Função: Ler arquivo (CSV/XLSX/XLS) e converter para “Parquet lógico” ─────
@st.cache_data(ttl=3600)
def load_and_convert_to_parquet(uploaded_file_bytes: bytes, file_name: str) -> dict[str, pd.DataFrame]:
    """
    Lê um arquivo CSV ou Excel e retorna um dicionário:
      - chave: nome da "aba" (para Excel) ou nome lógico (para CSV)
      - valor: DataFrame (já pronto para ser salvo em Parquet, se desejado)

    Aqui usamos o conceito de Parquet como formato interno ideal.
    Se quiser, é fácil também salvar fisicamente .parquet.
    """
    name = file_name.lower()
    sheets_data: dict[str, pd.DataFrame] = {}

    # CSV: tratamos como um dataset único
    if name.endswith(".csv"):
        # Para arquivos grandes, você pode usar parâmetros como:
        # dtype, usecols, etc.
        df = pd.read_csv(io.BytesIO(uploaded_file_bytes), dtype=str)
        df.columns = df.columns.str.strip()
        sheets_data["CSV_Dataset"] = df

    # Excel: múltiplas abas
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        xls = pd.ExcelFile(io.BytesIO(uploaded_file_bytes))
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
            df.columns = df.columns.str.strip()
            sheets_data[sheet_name] = df

    else:
        raise ValueError("Formato não suportado. Use .csv, .xlsx ou .xls")

    # Aqui, se quiser MESMO criar arquivos Parquet (por aba), dá pra fazer:
    # for sheet_name, df in sheets_data.items():
    #     buffer = io.BytesIO()
    #     df.to_parquet(buffer, index=False)
    #     buffer.seek(0)
    #     # Você poderia salvar esse buffer em disco ou em algum storage.
    #     # Neste app, vamos só manter os DataFrames em memória e tratar isso
    #     # como “equivalente lógico” ao Parquet.

    return sheets_data


# ── Popula o template Excel ───────────────────────────────────────────────────
def populate_template(df_data: pd.DataFrame, template_path: Path, column_mapping: dict) -> bytes | None:
    try:
        if not template_path.exists():
            raise FileNotFoundError(f"Template não encontrado em: {template_path}")

        template_wb = load_workbook(template_path)
        ws = template_wb.active

        template_cols = ['MATRICULA', 'TELEFONE', 'CONCESSIONARIA', 'CIDADE', 'DIRETORIA', 'SITUACAO']

        # Construímos as linhas do template a partir do df_data + mapping
        processed_rows_list = []
        for _, row_data_input in df_data.iterrows():
            new_row_dict = {}
            for t_col in template_cols:
                mapped_value = column_mapping.get(t_col)
                if mapped_value:
                    # Se mapped_value é uma coluna do df de entrada
                    if mapped_value in row_data_input.index:
                        new_row_dict[t_col] = row_data_input[mapped_value]
                    else:  # Valor fixo
                        new_row_dict[t_col] = mapped_value
                else:
                    new_row_dict[t_col] = ''
            processed_rows_list.append(new_row_dict)

        df_to_write = pd.DataFrame(processed_rows_list, columns=template_cols)

        # Escreve os dados no template a partir da linha 2 (linha 1 = cabeçalho do template)
        start_row = 2
        for r_idx, row_data in enumerate(df_to_write.itertuples(index=False), start=start_row):
            for c_idx, value in enumerate(row_data, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Salva o workbook em um buffer de bytes
        output_buffer = io.BytesIO()
        template_wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer.getvalue()

    except FileNotFoundError as e:
        st.error(
            f"Erro: {e}. Verifique se o arquivo '{TEMPLATE_FILE_NAME}' "
            "está na mesma pasta do aplicativo."
        )
        return None
    except Exception as e:
        st.error(f"Erro ao popular o template: {e}")
        return None


# ── Interface do Streamlit ────────────────────────────────────────────────────
st.title("Populador de Template Whats Infobip")
st.markdown("---")

# Verifica se o template existe
if not TEMPLATE_PATH.exists():
    st.error(
        f"**Erro:** O arquivo de template '{TEMPLATE_FILE_NAME}' não foi encontrado "
        "na mesma pasta do aplicativo."
    )
    st.info(
        "Por favor, coloque o arquivo `TEMPLATE_WHATS_COBRANCA.xlsx` "
        "ao lado do `app_populador_template_simples.py`."
    )
else:
    st.sidebar.header("1. Carregar Arquivo de Entrada")
    uploaded_file = st.sidebar.file_uploader(
        "Selecione um arquivo CSV ou Excel",
        type=["csv", "xlsx", "xls"],
    )

    sheets_data = None
    if uploaded_file:
        try:
            sheets_data = load_and_convert_to_parquet(
                uploaded_file.getvalue(),
                uploaded_file.name,
            )
            st.sidebar.success(
                f"Arquivo '{uploaded_file.name}' carregado e convertido (para uso tipo Parquet) com sucesso!"
            )
            st.sidebar.write(f"Abas/Datasets encontrados: {', '.join(sheets_data.keys())}")
        except Exception as e:
            st.sidebar.error(f"Falha ao carregar/convertar arquivo: {e}")
            sheets_data = None

    if sheets_data:
        # 2. Selecionar “abas” (ou dataset do CSV)
        st.sidebar.header("2. Selecionar Abas/Datasets para Processar")
        selected_sheets = st.sidebar.multiselect(
            "Escolha as abas/datasets que deseja processar:",
            options=list(sheets_data.keys()),
            default=list(sheets_data.keys()),
        )

        if selected_sheets:
            st.sidebar.header("3. Mapeamento de Colunas")
            st.sidebar.info(
                "Mapeie as colunas do seu arquivo (já convertido) "
                "para as colunas do template."
            )

            # Colunas do primeiro dataset selecionado
            first_selected_df_cols = ['']
            first_selected_df_cols.extend(
                sheets_data[selected_sheets[0]].columns.tolist()
            )

            template_cols = [
                'MATRICULA',
                'TELEFONE',
                'CONCESSIONARIA',
                'CIDADE',
                'DIRETORIA',
                'SITUACAO',
            ]

            column_mapping: dict[str, str] = {}

            for t_col in template_cols:
                mapping_type = st.sidebar.radio(
                    f"Como preencher '{t_col}'?",
                    ('Mapear Coluna', 'Valor Fixo'),
                    key=f"type_map_{t_col}",
                    index=0 if t_col not in ['CONCESSIONARIA'] else 1,
                )
                if mapping_type == 'Mapear Coluna':
                    selected_source_col = st.sidebar.selectbox(
                        f"Coluna do arquivo para '{t_col}':",
                        options=first_selected_df_cols,
                        key=f"map_{t_col}",
                    )
                    if selected_source_col:
                        column_mapping[t_col] = selected_source_col
                else:
                    fixed_value = st.sidebar.text_input(
                        f"Valor fixo para '{t_col}':",
                        value="" if t_col not in ['CONCESSIONARIA'] else f"Corsan",
                        key=f"fixed_val_{t_col}",
                    )
                    column_mapping[t_col] = fixed_value

            st.sidebar.markdown("---")
            st.sidebar.header("4. Gerar Arquivos XLSX")
            if st.sidebar.button("🚀 Processar e Gerar Arquivos"):
                if not TEMPLATE_PATH.exists():
                    st.error(
                        f"Erro: O arquivo de template '{TEMPLATE_FILE_NAME}' não foi encontrado."
                    )
                elif not column_mapping:
                    st.warning("Configure o mapeamento de colunas antes de processar.")
                else:
                    st.session_state["processed_files"] = {}
                    status_text = st.empty()
                    progress_bar = st.progress(0)

                    total = len(selected_sheets)
                    for i, sheet_name in enumerate(selected_sheets, start=1):
                        status_text.info(f"Processando dataset/aba: '{sheet_name}'...")
                        df_aba = sheets_data[sheet_name]

                        output_file_bytes = populate_template(
                            df_aba,
                            TEMPLATE_PATH,
                            column_mapping,
                        )

                        if output_file_bytes:
                            st.session_state["processed_files"][sheet_name] = output_file_bytes

                        progress_bar.progress(i / total)

                    status_text.success("Processamento concluído para todos os datasets/abas!")
                    st.rerun()

    # Download dos arquivos processados
    if "processed_files" in st.session_state and st.session_state["processed_files"]:
        st.subheader("Arquivos XLSX Gerados")
        st.markdown("Clique para baixar os arquivos gerados a partir do template:")
        for sheet_name, file_bytes in st.session_state["processed_files"].items():
            st.download_button(
                label=f"⬇️ Baixar {sheet_name}.xlsx",
                data=file_bytes,
                file_name=f"{sheet_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_{sheet_name}",
            )

    # Opção para limpar cache
    st.sidebar.markdown("---")
    st.sidebar.header("Opções de Cache")
    if st.sidebar.button("Limpar Cache de Dados Carregados"):
        st.cache_data.clear()
        st.sidebar.success("Cache de dados limpo. O app será recarregado.")
        st.rerun()
